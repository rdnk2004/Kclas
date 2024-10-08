import pandas as pd
import openpyxl
import os

# Load the Excel file
file_path = r'C:\EJS\Updated Front End\KCLAS\public\files\Book1.xlsx'
workbook = pd.read_excel(file_path, header=None)

# Determine the number of columns
num_columns = workbook.shape[1]

# Generate initial headers
headers = ['Register No', 'Name', 'College ID']
num_subjects = (num_columns - 3) // 4

for i in range(1, num_subjects + 1):
    headers.extend([f'Subject Code {i}', f'Subject Name {i}', f'Marks {i}', f'Result {i}'])

# Assign headers to the DataFrame
workbook.columns = headers

# Save the DataFrame to a new Excel file
temp_file_path = 'Processed_Semester_3.xlsx'
workbook.to_excel(temp_file_path, index=False, header=True)

# Load the saved workbook for further processing
workbook = openpyxl.load_workbook(temp_file_path)

# Function to split marks and calculate total
def process_marks(marks):
    if marks is None:
        return None, None, None
    if isinstance(marks, int):
        return marks, 0, marks
    marks = str(marks)
    if '+' not in marks:
        return None, None, None
    try:
        internal, external = map(int, marks.split('+'))
        total = internal + external
        return internal, external, total
    except ValueError:
        return None, None, None

# Select the first sheet
sheet = workbook.active

# Determine the starting column index for the first set of subject columns
start_col = 4

# Determine the number of subjects
num_subjects = (sheet.max_column - 3) // 4

# Iterate over each subject and insert Internal, External, and Total columns
for subject_num in range(num_subjects):
    # Calculate the position for inserting new columns
    marks_col = start_col + subject_num * 7 + 2
    insert_col = marks_col + 1

    # Insert new columns for Internal, External, and Total
    sheet.insert_cols(insert_col, 3)
    sheet.cell(row=1, column=insert_col, value=f'Internal {subject_num + 1}')
    sheet.cell(row=1, column=insert_col + 1, value=f'External {subject_num + 1}')
    sheet.cell(row=1, column=insert_col + 2, value=f'Total {subject_num + 1}')

    # Process marks for each row starting from the second row
    for row in range(2, sheet.max_row + 1):
        marks_cell_value = sheet.cell(row=row, column=marks_col).value
        internal, external, total = process_marks(marks_cell_value)
        
        # Update the cells with internal, external, and total marks
        sheet.cell(row=row, column=insert_col, value=internal)
        sheet.cell(row=row, column=insert_col + 1, value=external)
        sheet.cell(row=row, column=insert_col + 2, value=total)
        
for col in sheet.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2)
    sheet.column_dimensions[column].width = adjusted_width

# Update the headers after inserting new columns
updated_headers = [cell.value for cell in sheet[1]]

# Save the updated workbook
workbook.save(temp_file_path)

# Create new workbooks for each department
department_workbooks = {
    'bscdatascience_2022': openpyxl.Workbook(),
    'bscdatascience_2023': openpyxl.Workbook(),
    'bba_2021': openpyxl.Workbook(),
    'bba_2022': openpyxl.Workbook(),
    'bba_2023': openpyxl.Workbook(),
    'bbaib_2021': openpyxl.Workbook(),
    'bbaib_2022': openpyxl.Workbook(),
    'bbaib_2023': openpyxl.Workbook(),
    'bscpsychology_2021': openpyxl.Workbook(),
    'bscpsychology_2022': openpyxl.Workbook(),
    'bscpsychology_2023': openpyxl.Workbook(),
    'bcom_2021': openpyxl.Workbook(),
    'bcom_2022': openpyxl.Workbook(),
    'bcom_2023': openpyxl.Workbook(),
    'bcompa_2021': openpyxl.Workbook(),
    'bcompa_2022': openpyxl.Workbook(),
    'bcompa_2023': openpyxl.Workbook(),
    'batamilcreativewriting_2022': openpyxl.Workbook(),
    'batamilcreativewriting_2023': openpyxl.Workbook(),
    'bscvisualcommunication_2021': openpyxl.Workbook(),
    'bscvisual_communication_2022': openpyxl.Workbook(),
    'bscvisualcommunication_2023': openpyxl.Workbook(),
    'baeconomics_2021': openpyxl.Workbook(),
    'baeconomics_2022': openpyxl.Workbook(),
    'baeconomics_2023': openpyxl.Workbook(),
    'bapoliticalscience_2021': openpyxl.Workbook(),
    'bapoliticalscience_2022': openpyxl.Workbook(),
    'bapoliticalscience_2023': openpyxl.Workbook(),
    'mapoliticalscience_2022': openpyxl.Workbook(),
    'mapoliticalscience_2023': openpyxl.Workbook(),
    'msw_2022': openpyxl.Workbook(),
    'msw_2023': openpyxl.Workbook()
}

# Assign headers to each department workbook
for dept, wb in department_workbooks.items():
    ws = wb.active
    ws.append(updated_headers)

# Define department prefixes
department_prefixes = {
    'bscdatascience_2022': '2228M',
    'bscdatascience_2023': '2328M',
    'bba_2021': '2125F',
    'bba_2022': '2225F',
    'bba_2023': '2325F',
    'bbaib_2021': '2125N',
    'bbaib_2022': '2225N',
    'bbaib_2023': '2325N',
    'bcom_2021': '212AA',
    'bcom_2022': '222AA',
    'bcom_2023': '232AA',
    'bcompa_2021': '212AK',
    'bcompa_2022': '222AK',
    'bcompa_2023': '232AK',
    'bscpsychology_2021': '2126U',
    'bscpsychology_2022': '2226U',
    'bscpsychology_2023': '2326U',
    'bscvisualcommunication_2021': '2122S',
    'bscvisualcommunication_2022': '2222S',
    'bscvisualcommunication_2023': '2322S',
    'baeconomics_2021': '2121C',
    'baeconomics_2022': '2221C',
    'baeconomics_2023': '2321C',
    'batamilcreativewriting_2022': '2221G',
    'batamilcreativewriting_2023': '2321G',
    'msw_2022': '2231B',
    'msw_2023': '2331B',
    'bapoliticalscience_2021': '2121B',
    'bapoliticalscience_2022': '2221B',
    'bapoliticalscience_2023': '2321B',
    'mapoliticalscience_2022': '2231M',
    'mapoliticalscience_2023': '2331M'
}

# Process each row starting from the second row
for row in range(2, sheet.max_row + 1):
    register_no = sheet.cell(row=row, column=1).value
    row_values = [cell.value for cell in sheet[row]]
    
    for dept, prefix in department_prefixes.items():
        if register_no.startswith(prefix):
            dept_sheet = department_workbooks[dept].active
            dept_sheet.append(row_values)
            break

# Define output directory
output_directory = r'C:\EJS\Updated Front End\KCLAS\public\files'

# Save the department-specific workbooks
for dept, wb in department_workbooks.items():
    dept_file_path = os.path.join(output_directory, f'{dept.replace("_", " ").title().replace(" ", "_")}_dept.xlsx')
    wb.save(dept_file_path)

print('Students separated based on department and saved successfully!')
